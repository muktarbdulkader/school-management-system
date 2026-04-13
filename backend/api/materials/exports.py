"""
Data Export System - Exporters
Supports PDF, Excel, and CSV exports
"""
import io
import csv
import gc
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from django.http import HttpResponse
from django.template.loader import render_to_string


class PDFExporter:
    """PDF export functionality"""

    @staticmethod
    def export_report_card(student, term, grades, attendance_data):
        """Generate student report card PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph("STUDENT REPORT CARD", title_style))
        elements.append(Spacer(1, 0.3*inch))

        # Student Info
        student_info = [
            ['Student Name:', student.user.full_name],
            ['Grade:', student.grade.grade if student.grade else 'N/A'],
            ['Section:', student.section.name if student.section else 'N/A'],
            ['Term:', term.name],
            ['Academic Year:', term.academic_year],
        ]

        info_table = Table(student_info, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))

        # Grades Table
        elements.append(Paragraph("Academic Performance", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))

        grade_data = [['Subject', 'Grade', 'Remarks']]
        for grade in grades:
            grade_data.append([
                grade.subject_id.name if grade.subject_id else 'N/A',
                str(grade.score) if hasattr(grade, 'score') else (grade.grade if hasattr(grade, 'grade') else 'N/A'),
                grade.remarks if hasattr(grade, 'remarks') else 'Good'
            ])

        grade_table = Table(grade_data, colWidths=[3*inch, 1*inch, 2*inch])
        grade_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(grade_table)
        elements.append(Spacer(1, 0.3*inch))

        # Attendance
        elements.append(Paragraph("Attendance Summary", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))

        attendance_info = [
            ['Total Days:', str(attendance_data.get('total_days', 0))],
            ['Days Present:', str(attendance_data.get('days_present', 0))],
            ['Days Absent:', str(attendance_data.get('days_absent', 0))],
            ['Attendance %:', f"{attendance_data.get('percentage', 0):.2f}%"],
        ]

        attendance_table = Table(attendance_info, colWidths=[2*inch, 2*inch])
        attendance_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(attendance_table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Force garbage collection to free memory
        gc.collect()
        
        return buffer

    @staticmethod
    def export_list_to_pdf(title, headers, data):
        """Generic list export to PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 0.3*inch))

        # Table
        table_data = [headers] + data
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        
        # Force garbage collection to free memory
        gc.collect()
        
        return buffer


class ExcelExporter:
    """Excel export functionality"""

    @staticmethod
    def export_to_excel(title, headers, data, sheet_name='Sheet1'):
        """Generic Excel export"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Title
        ws.merge_cells('A1:' + get_column_letter(len(headers)) + '1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        ws.row_dimensions[1].height = 30

        # Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='e0e0e0', end_color='e0e0e0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Data
        for row_num, row_data in enumerate(data, 3):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Force garbage collection
        gc.collect()
        
        return buffer

    @staticmethod
    def export_attendance_sheet(class_name, students, dates):
        """Export attendance sheet template"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance'

        # Headers
        ws['A1'] = 'Student Name'
        ws['A1'].font = Font(bold=True)

        for idx, date in enumerate(dates, 2):
            cell = ws.cell(row=1, column=idx)
            cell.value = date.strftime('%Y-%m-%d')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', textRotation=45)

        # Students
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1).value = student.user.full_name

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Force garbage collection
        gc.collect()
        
        return buffer


class CSVExporter:
    """CSV export functionality"""

    @staticmethod
    def export_to_csv(headers, data):
        """Generic CSV export"""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(data)
        return buffer.getvalue()

    @staticmethod
    def export_queryset_to_csv(queryset, fields):
        """Export Django queryset to CSV"""
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Headers
        writer.writerow(fields)

        # Data
        for obj in queryset:
            row = [getattr(obj, field, '') for field in fields]
            writer.writerow(row)

        return buffer.getvalue()
